import locale, pyodbc
from datetime import datetime
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font
from django.http import HttpResponse
from django.shortcuts import render


# Définir la locale française
locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')

server = "172.16.69.1"
database = 'INTERCO'
username = 'reader'
password = 'm1234'
driver = "ODBC Driver 17 for SQL Server"

global_content = None
global_date_debut = None
global_date_fin = None

cnxn = pyodbc.connect(f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={username};PWD={password}")
cursor = cnxn.cursor()

    
def recap(request):
    date_debut = request.GET.get("start-date")
    date_fin = request.GET.get("end-date")

    societes_list = ["INVISO", "AGRIFARM", "AGRIFEED", "AGRIKOBA", "AGRIMOTORS", "AGRIVAL", "AGRIVET", "AGRIVET ANTSIRABE", "AGRIVET ANTSIRANANA", "AGRIVET FIANARANTSOA", "AGRIVETAM", "BLAST", "BOVIMA", "EUROPAINTS", "FARMANTSIKA", "FIRST ENERGY", "FLY LEASE", "FLY TECHNOLOGIES", "ID MOTORS", "ID RENTAL", "IDF", "INTERKEM", "INVISO DISTRIBUTION", "IOI SALAMA", "MABEL", "NUTRIFOOD", "ONG BOVIMA", "RYA", "SCIFD", "NOAH", "SIM", "SMTP", "UNITED MALAGASY"]
    context = {
        "societes_list": societes_list,
        "date_debut": date_debut,
        "date_fin": date_fin
    }

    global global_content

    # Exporter le tableau en Excel
    if request.GET.get("export") and global_content:
        return export_xls(global_content, societes_list)


    content = []
    if date_debut and date_fin:
        if cnxn:
            count = 0

            for x in societes_list: 
                count = count + 1
                
                for y in societes_list: 
                    val = getValues(x, y, date_debut, date_fin)
                    if val and len(val) > 0 and val[0] is not None:
                        try:
                            formatted_value = locale.format_string('%.2f', float(val[0]), grouping=True)
                        except ValueError as ve:
                            print(f"Erreur de formatage pour {x} et {y}: {ve}")
                            formatted_value = "0,00"
                    else:
                        formatted_value = "0,00"

                    resultat = {
                        'societex': x,
                        'societey': y,  
                        'valeur': formatted_value
                    }
                    content.append(resultat)

            context["content"] = content
            global_content = content

    return render(request, 'recap/recap.html', context)



def export_xls(content, societes_list):
    # Préparer les données pour le DataFrame
    data = []
    for societe in societes_list:
        row = [societe]
        for societe2 in societes_list:
            if societe == societe2:
                row.append("")  # Case vide pour la diagonale
            else:
                # Récupérer la valeur et convertir en float si possible
                val = next((item['valeur'] for item in content if item['societex'] == societe and item['societey'] == societe2), "0,00")
                try:
                    # Convertir en float après nettoyage des chaînes
                    val = float(val.replace(",", ".").replace(" ", "").replace("\xa0", ""))
                except ValueError:
                    val = 0.0  # Valeur par défaut si la conversion échoue
                row.append(val)
        data.append(row)

    # Création du DataFrame
    df = pd.DataFrame(data, columns=[""] + societes_list)

    # Exportation en Excel avec `pandas`
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, sheet_name='Intercompany', index=False)

    # Personnalisation des styles (background et texte)
    sheet = writer.sheets['Intercompany']

    # Définir les styles
    blue_fill = PatternFill(start_color="005ec2", end_color="005ec2", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    yellow_fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")  # Couleur jaune pour la diagonale

    # Appliquer le style aux colonnes (header)
    for cell in sheet[1]:
        cell.fill = blue_fill
        cell.font = white_font

    # Appliquer le style à la première colonne (les sociétés)
    for row in sheet.iter_rows(min_row=2, max_row=len(societes_list) + 1, min_col=1, max_col=1):
        for cell in row:
            cell.fill = blue_fill
            cell.font = white_font

    # Colorier les cellules diagonales en jaune
    for i in range(2, len(societes_list) + 2):
        sheet.cell(row=i, column=i).fill = yellow_fill

    # Ajuster la largeur des colonnes en fonction du contenu (y compris l'en-tête)
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        
        # Vérifier chaque cellule de la colonne
        for cell in col:
            if cell.value:
                # Vérifier la longueur de la chaîne et ajuster la longueur maximale
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
        
        adjusted_width = max_length + 2  # Ajouter un peu d'espace pour l'esthétique
        sheet.column_dimensions[column].width = adjusted_width

    # Appliquer le format de nombre aux cellules contenant des valeurs numériques
    number_format = '#,##0.00'  # Format numérique avec deux décimales
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=sheet.max_column):
        for cell in row:
            if isinstance(cell.value, (float, int)):  # Vérifier si c'est une valeur numérique
                cell.number_format = number_format

    writer.close()

    # Préparer la réponse HTTP
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="recap.xlsx"'

    return response


def getValues(societe, tiers, date_start, date_end):
    month_start = None
    month_end = None
    year_start = None
    year_end = None

    if date_start:
        start_obj = datetime.strptime(date_start, '%Y-%m')
        year_start = start_obj.year
        month_start = start_obj.month

    if date_end:
        end_obj = datetime.strptime(date_end, '%Y-%m')
        year_end = end_obj.year
        month_end = end_obj.month

    if tiers and societe and month_start and month_end and year_start and year_end:
        try:
            cursor.execute(f""" SELECT 
                                    SUM(isnull(GRAND_LIVRE.SOLDE, 0)) AS SOLDE
                                FROM 
                                    GRAND_LIVRE
                                JOIN 
                                    TABLE_TIERS 
                                    ON GRAND_LIVRE.TIERS = TABLE_TIERS.CODE OR GRAND_LIVRE.COMPTE = TABLE_TIERS.COMPTE_GENERAL
                                WHERE 
                                    TABLE_TIERS.TIERS = '{tiers}' 
                                    AND TABLE_TIERS.SOCIETE = '{societe}' 
                                    AND GRAND_LIVRE.SOCIETE = '{societe}'
                                    AND month(DATE_COMPTABLE) BETWEEN '{month_start}' AND '{month_end}' AND year(DATE_COMPTABLE) BETWEEN '{year_start}' AND '{year_end}'
                                GROUP BY GRAND_LIVRE.SOCIETE, TABLE_TIERS.TIERS """)
            return cursor.fetchone()

        except pyodbc.Error as e:
            print("Erreur: ", e)




