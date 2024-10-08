from django.shortcuts import render
import locale, pyodbc
from datetime import datetime
from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment


# Définir la locale française
locale.setlocale(locale.LC_ALL, 'fr_FR.UTF-8')

server = "172.16.69.1"
database = 'INTERCO'
username = 'reader'
password = 'm1234'
driver = "ODBC Driver 17 for SQL Server"


cnxn = pyodbc.connect(f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={username};PWD={password}")
cursor = cnxn.cursor()

global_content = None
global_total_values = None
global_data_pop = None


def detail(request):
    date_debut = request.POST.get("start-date")
    date_fin = request.POST.get("end-date")
    societe_select = request.POST.get("societe_select")

    export_pop = request.GET.get("export_pop")


    if export_pop:
        return export_popup_to_xls()
    

    if date_debut:
        start_obj = datetime.strptime(date_debut, '%Y-%m')
        year_start = start_obj.year
        month_start = start_obj.month

    if date_fin:
        end_obj = datetime.strptime(date_fin, '%Y-%m')
        year_end = end_obj.year
        month_end = end_obj.month


    global global_content
    global global_total_values

    headers = ["TIERS", "INTERCO GLOBAL", "INTERCO FINANCIER", "INTERCO CompteCourant", "INTERCO COMMERCIAL", "INTERCO ND-REFACTURATION", "AUTRE"]

    societes_list = ["INVISO", "AGRIFARM", "AGRIFEED", "AGRIKOBA", "AGRIMOTORS", "AGRIVAL", "AGRIVET", "AGRIVET ANTSIRABE", "AGRIVET ANTSIRANANA", "AGRIVET FIANARANTSOA", "AGRIVETAM", "BLAST", "BOVIMA", "EUROPAINTS", "FARMANTSIKA", "FIRST ENERGY", "FLY LEASE", "FLY TECHNOLOGIES", "ID MOTORS", "ID RENTAL", "IDF", "INTERKEM", "INVISO DISTRIBUTION", "IOI SALAMA", "MABEL", "NUTRIFOOD", "ONG BOVIMA", "RYA", "SCIFD", "NOAH", "SIM", "SMTP", "UNITED MALAGASY"]

    context = {
        "societes_list": societes_list,
        "societe_select": societe_select,
        "date_debut": date_debut,
        "date_fin": date_fin
    }

    type_interco = ["global", "Financier", "Compte Courant", "Commercial", "ND - Refac", "Autre"]
    type_interco_ttl = ["global", "Financier", "CompteCourant", "Commercial", "NDRefac", "Autre"]
    # totals = {type: 0 for type in type_interco}
    totals = {type: 0 for type in type_interco_ttl}


    # Exporter le tableau en Excel
    if request.POST.get("export") and global_content and global_total_values:
        return export_xls(global_content, societes_list, type_interco, global_total_values, headers)
    



    if societe_select and year_start and year_end and month_start and month_end:
        if cnxn:
            content = []

            for tiers in societes_list:
                # for type in type_interco:
                for type in type_interco:
                    # print("type: ", type)

                    val = getValues(societe_select, tiers, type, year_start, year_end, month_start, month_end)
                    valeur = float(val[0]) if val else 0.0
                    resultat = {
                        'tiers': tiers,
                        'type_interco': type,
                        'valeur': locale.format_string('%.2f', valeur, grouping=True) if valeur else "0,00"
                    }
                    content.append(resultat)
                    type = type.replace("-", "").replace(" ", "")
                    totals[type] += valeur


            total_values = {k: locale.format_string('%.2f', v, grouping=True) for k, v in totals.items()}
            context["content"] = content
            context["type_interco"] = type_interco
            context["tiers_list"] = societes_list
            context["total_values"] = total_values

            global_content = content
            global_total_values = total_values



    return render(request, 'detail/detail.html', context)



def export_xls(content, societes_list, type_interco, total_values, headers):
    # Créer une DataFrame avec des valeurs numériques
    data = []
    for tiers in societes_list:
        row = [tiers]  # La première colonne contient les tiers
        for type in type_interco:
            for item in content:
                if item['tiers'] == tiers and item['type_interco'] == type:
                    valeur = item['valeur']
                    if isinstance(valeur, str):
                        # Supprimer les espaces insécables et remplacer les virgules
                        valeur = valeur.replace('\xa0', '').replace(',', '')
                        try:
                            valeur = float(valeur)  # Essayer de convertir en float
                        except ValueError:
                            valeur = None  # Gérer les valeurs qui ne peuvent pas être converties
                    row.append(valeur)
        data.append(row)

    # Ajouter une ligne total avec des valeurs numériques
    total_row = ["TOTAL"]
    for type in type_interco:
        type = type.replace("-", "").replace(" ", "")
        total_value = total_values.get(type, None)  # Utiliser get pour éviter KeyError
        if isinstance(total_value, str):
            # Supprimer les espaces insécables et remplacer les virgules
            total_value = total_value.replace('\xa0', '').replace(',', '')
            try:
                total_value = float(total_value)
            except ValueError:
                total_value = None
        total_row.append(total_value)
    data.append(total_row)

    # Créer un DataFrame avec les données
    df = pd.DataFrame(data, columns=["TIERS"] + headers[1:])

    # Créer un writer Excel en utilisant openpyxl
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Data')

        # Obtenir la feuille de travail
        worksheet = writer.sheets['Data']

        # Ajuster la largeur des colonnes pour s'adapter au contenu
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Obtenir le nom de la colonne
            for cell in col:
                try:
                    if cell.value:  # Vérifier si la cellule a une valeur
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2  # Ajouter un peu de marge
            worksheet.column_dimensions[column].width = adjusted_width

        # Appliquer le format de nombre aux colonnes contenant des valeurs numériques
        number_format = '#,##0.00'  # Format numérique avec milliers et deux décimales
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
            for cell in row:
                if isinstance(cell.value, (float, int)):  # Vérifier si c'est une valeur numérique
                    cell.number_format = number_format

        # Style des en-têtes : fond bleu #005ec2, texte blanc, gras
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="005ec2", end_color="005ec2", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")

        for cell in worksheet[1]:  # Première ligne (en-têtes)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # Style de la première colonne (colonne des tiers)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="005ec2", end_color="005ec2", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Colorier la première cellule de la dernière ligne (TOTAL) en bleu clair
        total_cell = worksheet.cell(row=worksheet.max_row, column=1)  # Première colonne dans la dernière ligne
        total_cell.font = Font(bold=True, color="000000")
        total_cell.fill = PatternFill(start_color="95fff6", end_color="95fff6", fill_type="solid")

        # Appliquer le format de nombre à toute la ligne du total
        for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row, min_col=2, max_col=worksheet.max_column):
            for cell_value in cell:
                if isinstance(cell_value.value, (float, int)):
                    cell_value.number_format = number_format

    # Réponse HTTP pour télécharger le fichier
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="details_par_société.xlsx"'
    response.write(output.getvalue())
    return response


def getValues(societe, tiers, type, year_start, year_end, month_start, month_end):

    if type == "global":
        type_interco = ('Autre', 'Commercial', 'Compte Courant', 'Financier', 'ND - Refac')
    else:
        type_interco = "('" + type + "')"


    try:
        req = f""" SELECT 
                                    SUM(isnull(GRAND_LIVRE.SOLDE, 0)) AS SOLDE
                                FROM 
                                    GRAND_LIVRE
                                JOIN 
                                    TABLE_TIERS 
                                    ON GRAND_LIVRE.TIERS = TABLE_TIERS.CODE OR GRAND_LIVRE.COMPTE = TABLE_TIERS.COMPTE_GENERAL
                                WHERE 
                                    TABLE_TIERS.TIERS = '{tiers}' 
                                    AND TABLE_TIERS.SOCIETE = '{societe}' 
                                    AND GRAND_LIVRE.SOCIETE = '{societe}'
                                    AND GRAND_LIVRE.TYPE_INTERCO IN {type_interco}
                                    AND month(GRAND_LIVRE.DATE_COMPTABLE) BETWEEN '{month_start}' AND '{month_end}' AND year(GRAND_LIVRE.DATE_COMPTABLE) BETWEEN '{year_start}' AND '{year_end}'  
                                GROUP BY GRAND_LIVRE.SOCIETE, TABLE_TIERS.TIERS """
        


        cursor.execute(req)
        return cursor.fetchone()

    except pyodbc.Error as e:
        print("Erreur: ", e)




def getPopupValues(societe, tiers, type_interco, year_start, year_end, month_start, month_end):

    type_filtre = f" AND GRAND_LIVRE.TYPE_INTERCO IN ('{type_interco}') "

    if type_interco == "global":
        type_filtre = ""
        

    try:

        req = f""" SELECT COMPTE, INTITULE_COMPTE, DATE_COMPTABLE, DATE_SAISIE, JOURNAL, PIECE, FACTURE, LIBELLE, GRAND_LIVRE.TIERS, INTITULE_TIERS, ECHEANCE, LETTRAGE, DEBIT, CREDIT, TYPE_LETTRAGE, SOLDE, INTERCO, ANNEE_MOIS, TYPE_INTERCO                                        
            FROM 
                GRAND_LIVRE
            JOIN 
                TABLE_TIERS 
                ON GRAND_LIVRE.TIERS = TABLE_TIERS.CODE OR GRAND_LIVRE.COMPTE = TABLE_TIERS.COMPTE_GENERAL
            WHERE 
                TABLE_TIERS.TIERS = '{tiers}' 
                AND TABLE_TIERS.SOCIETE = '{societe}' 
                AND GRAND_LIVRE.SOCIETE = '{societe}'
                {type_filtre}
                AND month(GRAND_LIVRE.DATE_COMPTABLE) BETWEEN '{month_start}' AND '{month_end}' AND year(GRAND_LIVRE.DATE_COMPTABLE) BETWEEN '{year_start}' AND '{year_end}'  
                """
        
        print("req: ", req)

        cursor.execute(req)
        return cursor.fetchall()

    except pyodbc.Error as e:
        print("Erreur: ", e)
        return None





def detail_Individuel(request):

    societe_select = request.GET.get("societe_select")
    date_debut = request.GET.get("start")
    date_fin = request.GET.get("end")
    tiers = request.GET.get("tiers")
    typex = request.GET.get("type")

    global global_data_pop
    
    year_start = None
    year_end = None
    month_start = None
    month_end = None

    if date_debut:
        start_obj = datetime.strptime(date_debut, '%Y-%m')
        year_start = start_obj.year
        month_start = start_obj.month

    if date_fin:
        end_obj = datetime.strptime(date_fin, '%Y-%m')
        year_end = end_obj.year
        month_end = end_obj.month



    if cnxn and societe_select and date_debut and date_fin and year_start and year_end and month_start and month_end:

        data = getPopupValues(societe_select, tiers, typex, year_start, year_end, month_start, month_end)

        if data:

            lines = []
            total_debit = 0
            total_credit = 0
            total_solde = 0
            for row in data:
                lines.append({       
                    'COMPTE': row[0], 'INTITULE_COMPTE': row[1], 'DATE_COMPTABLE': row[2].strftime("%d/%m/%Y"), 'DATE_SAISIE': row[3].strftime("%d/%m/%Y"), 'JOURNAL': row[4], 'PIECE': row[5], 'FACTURE': row[6], 'LIBELLE': row[7], 'TIERS': row[8], 'INTITULE_TIERS': row[9], 'ECHEANCE': row[10].strftime("%d/%m/%Y"), 'LETTRAGE': row[11], 'DEBIT': row[12], 'CREDIT': row[13], 'TYPE_LETTRAGE': row[14], 'SOLDE': row[15], 'INTERCO': row[16], 'ANNEE_MOIS': row[17], 'TYPE_INTERCO': row[18]
                })
                total_debit += float(row[12])
                total_credit += float(row[13])
                total_solde += float(row[15])


            # add total
            lines.append({       
                'COMPTE': "TOTAL", 'INTITULE_COMPTE': "", 'DATE_COMPTABLE': "", 'DATE_SAISIE': "", 'JOURNAL': "", 'PIECE': "", 'FACTURE': "", 'LIBELLE': "", 'TIERS': "", 'INTITULE_TIERS': "", 'ECHEANCE': "", 'LETTRAGE': "", 'DEBIT': total_debit, 'CREDIT': total_credit, 'TYPE_LETTRAGE': "", 'SOLDE': total_solde, 'INTERCO': "", 'ANNEE_MOIS': "", 'TYPE_INTERCO': ""
            })

            resultat = {
                "data": lines,
            }

            global_data_pop = lines

            # print()
            # print(global_data_pop)
            # print()

            # Return the result as JSON
            return JsonResponse(resultat)
        
        else:
            print("eeeeeeeeeeeeeeeeeeeeeeeeee")

  



def export_popup_to_xls():
    global global_data_pop

    # Define the headers
    labels = [
        "COMPTE", "INTITULE_COMPTE", "DATE_COMPTABLE", "DATE_SAISIE", "JOURNAL", 
        "PIECE", "FACTURE", "LIBELLE", "TIERS", "INTITULE_TIERS", "ECHEANCE", 
        "LETTRAGE", "DEBIT", "CREDIT", "TYPE_LETTRAGE", "SOLDE", "INTERCO", 
        "ANNEE_MOIS", "TYPE_INTERCO"
    ]
    
    # Convert global_data_pop to a DataFrame
    df = pd.DataFrame(global_data_pop, columns=labels)

    # Convert "DEBIT", "CREDIT", and "SOLDE" columns to numeric values
    df["DEBIT"] = pd.to_numeric(df["DEBIT"].astype(str).str.replace('\xa0', '').str.replace(',', ''), errors='coerce')
    df["CREDIT"] = pd.to_numeric(df["CREDIT"].astype(str).str.replace('\xa0', '').str.replace(',', ''), errors='coerce')
    df["SOLDE"] = pd.to_numeric(df["SOLDE"].astype(str).str.replace('\xa0', '').str.replace(',', ''), errors='coerce')

    # Create an HTTP response with XLS content type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="details_pop.xlsx"'

    # Create an Excel writer using openpyxl
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')

        # Access the workbook and sheet for styling
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Define styles for the header row (first row)
        header_fill = PatternFill(start_color="005ec2", end_color="005ec2", fill_type="solid")  # Blue background
        header_font = Font(color="FFFFFF", bold=True)  # White bold text

        # Apply styles to the header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Get the last row number (since DataFrame starts without headers, pandas starts from row 2)
        last_row = len(df) + 1

        # Apply green background to the first cell of the last row
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green background
        worksheet[f'A{last_row}'].fill = green_fill

        # Apply bold to the entire last row
        bold_font = Font(bold=True)
        for cell in worksheet[last_row]:
            cell.font = bold_font

        # Format M, N, P columns (12th, 13th, and 15th columns)
        millennium_format = '#,##0.00'
        for row in worksheet.iter_rows(min_row=2, max_row=last_row, min_col=13, max_col=13):
            for cell in row:
                cell.number_format = millennium_format  # DEBIT column

        for row in worksheet.iter_rows(min_row=2, max_row=last_row, min_col=14, max_col=14):
            for cell in row:
                cell.number_format = millennium_format  # CREDIT column

        for row in worksheet.iter_rows(min_row=2, max_row=last_row, min_col=16, max_col=16):
            for cell in row:
                cell.number_format = millennium_format  # SOLDE column

        # **Auto-size all columns**
        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if cell.value:
                        # Adjust max length to handle large numeric values
                        max_length = max(max_length, len(str(cell.value)) + 2)
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add a little extra space for padding
            worksheet.column_dimensions[column].width = adjusted_width

    return response
