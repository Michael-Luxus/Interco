from django.shortcuts import render
import pyodbc
from dotenv import load_dotenv
load_dotenv()
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import Workbook
import pandas as pd
from io import BytesIO

server = "172.16.69.1"
database = 'INTERCO'
username = 'reader'
password = 'm1234'
driver = "ODBC Driver 17 for SQL Server"

global_result = None

cnxn = pyodbc.connect(f"DRIVER={{{driver}}};SERVER={server};DATABASE={database};UID={username};PWD={password}")
cursor = cnxn.cursor()


@login_required
def tiers(request):
    if request.POST.get("export"):
        return export_simple_xls()

    cursor.execute("SELECT DISTINCT SOCIETE AS SOCNAME FROM TABLE_TIERS ORDER BY SOCIETE ASC")
    societes = cursor.fetchall()

    context = {
        'societes': societes,
    }
    return render(request, 'tiers/tiers.html', context)


def export_simple_xls():
    global global_result

     # Create DataFrame from the data
    df = pd.DataFrame(global_result)

    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define headers
    headers = ['Code', 'Intitulé', 'Tiers', 'Compte']
    ws.append(headers)  # Write headers

    # Apply styling to header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="005ec2", end_color="005ec2", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    # Add data to the sheet (adjusted to exclude 'id' and include 'code', 'intitule', 'tiers', 'compte')
    for _, row in df.iterrows():
        ws.append([row['code'], row['intitule'], row['tiers'], row['compte']])

    # Adjust column widths to fit content
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Save the workbook to an HttpResponse object
    response = HttpResponse(content_type="application/ms-excel")
    response['Content-Disposition'] = 'attachment; filename="tiers.xlsx"'
    with BytesIO() as b:
        wb.save(b)
        response.write(b.getvalue())
    return response


def listTiers(request):
    societe = request.GET.get('societe', None)
    query = "SELECT * FROM TABLE_TIERS WHERE SOCIETE = ? ORDER BY CODE ASC"
    
    global global_result

    cursor.execute(query, (societe,)) 
    data = cursor.fetchall()

    result = []
    for row in data:
        result.append({       
            'code': row[1],     
            'intitule': row[2],   
            'tiers': row[3],  
            'compte': row[4],
            'id': row[5]
        })
        global_result = result
    return JsonResponse({'data': result, 'status': 'success'})

    
def modif_Tiers(request):
    global global_result
    id = request.GET.get('id')

    code = request.GET.get('code', '')
    if code.strip() == '':
        code = ' '

    intitule = request.GET.get('intitule', '')
    if intitule.strip() == '':
        intitule = ' '

    tiers = request.GET.get('tiers', '')
    if tiers.strip() == '':
        tiers = ' '

    compteg = request.GET.get('compteg') 
    if compteg.strip() == '':
        compteg = ' '

    societe = request.GET.get('societe', '')

    query = "UPDATE TABLE_TIERS SET CODE=?, INTITULE=? , TIERS=? , COMPTE_GENERAL=? WHERE ID = ?"
    
    if id and code and intitule and tiers and compteg and societe:
        cursor.execute(query, (code, intitule, tiers, compteg, id)) 
        print("Update successful")        
    else:
        print("KO")
    

    query2 = "SELECT * FROM TABLE_TIERS WHERE SOCIETE = ? ORDER BY CODE ASC"
    cursor.execute(query2, (societe,)) 
    data = cursor.fetchall()

    result = []
    for row in data:
        result.append({       
            'code': row[1],     
            'intitule': row[2],   
            'tiers': row[3],  
            'compte': row[4],
            'id': row[5]
        })
    
    global_result = result

    return JsonResponse({'data': result, 'status': 'success'})
    
